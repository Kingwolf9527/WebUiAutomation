# - * - coding:utf-8
# __author__ : kingwolf
# createtime : 2021/8/17 3:46

from openpyxl import load_workbook
from common.common_convert_data_type import CommonConvertData
from common import common_file_path
from common.common_log import CommonLog

logger = CommonLog.get_logger()


class CommonExcel(object):

    def __init__(self, file_path=None, sheet_name=None):
        """
        初始化Excel文件
        @param file_path:
        @param sheet_name:
        """
        if file_path is None:
            self.file_path = common_file_path.default_excel_path
        else:
            self.file_path = file_path

        if sheet_name is None:
            self.sheet_name = common_file_path.default_sheet_name
        else:
            self.sheet_name = sheet_name

        logger.info(f"------- 当前打开的Excel的路径为：{file_path} -------")
        # 创建workbook(workbook)对象
        self.workbook = load_workbook(filename=self.file_path)
        logger.info(f"------- 创建workbook对象 -------")
        # 创建sheet(worksheet)对象
        self.sheet = self.get_sheet()

    def get_sheet(self):
        """
        获取目标sheet
        @return:
        """
        target_sheet = self.workbook[self.sheet_name]
        return target_sheet

    def get_max_column(self):
        """
        获取target_sheet最大列
        @return:
        """
        max_col = self.sheet.max_column
        if max_col:
            logger.info(f"------- 当前sheet的最大列数为：{max_col} -------")
            return max_col
        else:
            logger.error("\r\n" + "------- 获取当前sheet最大列数据失败！ -------")

    def get_max_row(self):
        """
        获取target_sheet最大行
        @return:
        """
        max_row = self.sheet.max_row
        if max_row:
            logger.info(f"------- 当前sheet的最大行数为：{max_row} -------")
            return max_row
        else:
            logger.error("\r\n" + "------- 获取当前sheet最大行数据失败！ -------")

    def get_cell_value(self, row, col):
        """
        获取单元格的值
        @param row: row跟col索引从1开始
        @param col:
        @return:
        """
        if 1 <= row <= self.get_max_row() and 1 <= col <= self.get_max_column():
            cell_value = self.sheet.cell(row, col).value
            return cell_value
        else:
            logger.error("------- 请检查输入的row或者col值【row或者col至少从1开始，也不能超过最大值！】-------")

    def get_row_datas(self, row):
        """
        根据行号，获取一整行的数据
        @param row:
        @return:
        """
        row_datas = []
        for i in range(1, self.get_max_column() + 1):
            cell_value = self.get_cell_value(row=row, col=i)
            row_datas.append(cell_value)
        return row_datas

    def get_col_datas(self, col):
        """
        根据列号，获取一整列的数据
        @param col:
        @return:
        """
        col_datas = []
        for i in range(1, self.get_max_row() + 1):
            cell_value = self.get_cell_value(row=i, col=col)
            col_datas.append(cell_value)
        return col_datas

    def get_row_number(self, case_id):
        """
        根据测试用例的case_id来获取目标行的内容
        @param case_id: 测试用例的id
        @return:
        """
        # openpyxl的起始值为1
        defalut_number = 1
        # 默认case_id在用例的第一列
        case_id_values = self.get_col_datas(col=1)

        for case_id_value in case_id_values:
            if case_id == case_id_value:
                logger.info(f'------- 当前case_id为：{case_id}，对应第 {defalut_number} 行数据 -------')
                return defalut_number
            # 递归处理
            defalut_number += 1
        else:
            logger.error(f"------- 当前的：{case_id}有误，请确认！ -------")

    def get_row_datas_for_number(self, case_id):
        """
        根据对应的case_id获取对应行的内容，
        @param case_id:
        @return:
        """
        row_number = self.get_row_number(case_id)
        if row_number:
            row_datas = self.get_row_datas(row_number)
            return row_datas

    def dispose_datas(self):
        """
        组合处理数据
        @return:
        """
        # 存储处理好的数据
        case_datas = []
        for i in range(2, self.get_max_row() + 1):
            # 处理单元格的整形数据或者浮点型数据
            case_data = [CommonConvertData.fload_int_to_string(target=j) for j in self.get_row_datas(i)]
            logger.info(f"------- 第{i}行的数据为：{case_data} -------")
            case_datas.append(case_data)
        return case_datas

    def write_datas(self, row, col, value=None):
        """
        往Excel中写入数据
        @param row:
        @param col:
        @param value:
        @return:
        """
        try:
            if 1 <= row <= self.get_max_row() and 1 <= col <= self.get_max_column():
                self.sheet.cell(row, col).value = value
                logger.info(f"------- 正在向excel中的当前sheet：{self.sheet_name} \
                            第{row}，第{col}列，写入数据：{value} -------")
            else:
                logger.error(f"------- 写入数据失败！请检查输入的sheet_name，row，col参数是否有误！ -------")
        except Exception as e:
            self.sheet.cell(row, col).value = e
        finally:
            # 保存数据，退出
            self.workbook.save(self.file_path)
            logger.info(f"------- 写入数据成功，正在保存！文件保存路径为：{self.file_path} -------")

if __name__ == '__main__':
    dd = CommonExcel()